import openpyxl
import os

file_path = r"c:\Users\Titan\Documents\TU-Darmstadt\2025_SoSe\MASTER THESIS\Latex Thesis\tables\Digital Services Price Index (DSPI).xlsx"

print(f"Auditing file for charts: {file_path}")

try:
    # Load workbook with data_only=False to preserve charts (though openpyxl doesn't fully support reading all chart types, it might detect them)
    # Actually, openpyxl supports reading charts.
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)
    
    chart_count = 0
    sheets_with_charts = []

    print(f"Total Sheets: {len(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Check for charts
        # openpyxl stores charts in ws._charts usually, or ws.charts (depending on version)
        # Let's try both or inspect.
        charts = getattr(ws, 'charts', []) or getattr(ws, '_charts', [])
        
        if charts:
            print(f"\n[SHEET] {sheet_name}: Found {len(charts)} chart(s)")
            for i, chart in enumerate(charts):
                title = getattr(chart, 'title', "No Title")
                # Sometimes title is an object or None
                if title and not isinstance(title, str):
                    try:
                        title = title.tx.rich.p[0].r[0].t
                    except:
                        title = str(title)
                
                print(f"  - Chart {i+1}: {title} (Type: {type(chart).__name__})")
                sheets_with_charts.append((sheet_name, title))
            chart_count += len(charts)
        else:
            # Check for Drawing objects which might contain charts
            if hasattr(ws, 'drawing') and ws.drawing:
                 print(f"[SHEET] {sheet_name}: Has drawing object (potential chart/image)")
            else:
                 pass # print(f"[SHEET] {sheet_name}: No charts found")

    print(f"\nTotal Charts Found: {chart_count}")
    
    if chart_count == 0:
        print("WARNING: No chart objects detected. This might mean the visualizations are images, or openpyxl cannot see them, or they are not in this file.")

except Exception as e:
    print(f"Error auditing Excel file: {e}")
